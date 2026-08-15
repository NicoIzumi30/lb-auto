import json
import textwrap
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import urljoin

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont, ImageOps

from .database import row_to_dict


RED = "EC1D23"
DARK = "17201E"
MUTED = "6F7976"
LIGHT = "F4F5F2"
PHOTO_LABELS = ["Depan", "Belakang", "Sisi kanan", "Sisi kiri", "Ruang mesin", "Interior", "Nomor rangka", "Odometer"]


def _row(connection, table, unit_id):
    return row_to_dict(connection.execute(f"SELECT * FROM {table} WHERE unit_id=?", (unit_id,)).fetchone())


def collect_report_data(connection):
    units = []
    for raw in connection.execute("SELECT * FROM units ORDER BY created_at DESC"):
        unit = row_to_dict(raw)
        for table, key in (
            ("initial_qc", "initial_qc"), ("legal_prechecks", "legal_precheck"),
            ("inspections", "inspection"), ("payments", "payment"), ("repairs", "repair"),
            ("documents", "documents"), ("listings", "listing"), ("sales", "sale"),
            ("greetings", "greeting"),
        ):
            unit[key] = _row(connection, table, unit["id"])
        unit["audit"] = [row_to_dict(row) for row in connection.execute(
            """SELECT a.*,u.name user_name FROM audit_logs a LEFT JOIN users u ON u.id=a.user_id
               WHERE a.entity_id=? ORDER BY a.created_at""", (unit["id"],)
        )]
        hpp = (unit.get("buy_price") or 0) + (unit.get("repair_cost") or 0)
        unit["hpp"] = hpp
        unit["profit"] = (unit.get("sell_price") or 0) - hpp
        units.append(unit)
    sold = [unit for unit in units if unit["status"] == "SOLD_DELIVERED"]
    return {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "units": units,
        "summary": {
            "total_units": len(units), "sold_units": len(sold),
            "revenue": sum(unit.get("sell_price") or 0 for unit in sold),
            "hpp": sum(unit["hpp"] for unit in sold),
            "profit": sum(unit["profit"] for unit in sold),
        },
    }


def _list(value):
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    except (TypeError, json.JSONDecodeError):
        return []


def _url(base_url, value):
    if not value:
        return ""
    value = str(value)
    if value.startswith(("http://", "https://")):
        return value
    return urljoin(base_url, value.lstrip("/"))


def unit_photos(unit, base_url):
    photos = []
    for index, value in enumerate(_list(unit.get("source_photos"))):
        photos.append(("Cover", f"Cover {index + 1}", _url(base_url, value), value))
    inspection = unit.get("inspection") or {}
    for index, value in enumerate(_list(inspection.get("photos"))):
        photos.append(("Inspeksi", PHOTO_LABELS[index] if index < len(PHOTO_LABELS) else f"Foto {index + 1}", _url(base_url, value), value))
    payment = unit.get("payment") or {}
    if payment.get("proof_url"):
        photos.append(("Pembayaran", "Bukti pembayaran", _url(base_url, payment["proof_url"]), payment["proof_url"]))
    repair = unit.get("repair") or {}
    for index, value in enumerate(_list(repair.get("before_photos"))):
        photos.append(("Repair before", f"Before {index + 1}", _url(base_url, value), value))
    for index, value in enumerate(_list(repair.get("after_photos"))):
        photos.append(("Repair after", f"After {index + 1}", _url(base_url, value), value))
    listing = unit.get("listing") or {}
    for index, value in enumerate(_list(listing.get("media_items"))):
        photos.append(("Listing", f"Media {index + 1}", _url(base_url, value), value))
    greeting = unit.get("greeting") or {}
    if greeting.get("media_url"):
        photos.append(("Greeting", "Delivery greeting", _url(base_url, greeting["media_url"]), greeting["media_url"]))
    return photos


def _sheet(workbook, title, headers):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=RED)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return sheet


def _finish_sheet(sheet, currency_columns=()):
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 100) + 1)]
        width = min(55, max(11, max(map(len, values), default=11) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width
    for column in currency_columns:
        for cell in sheet[column][1:]:
            cell.number_format = '"Rp" #,##0'
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_excel(data, base_url):
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Ringkasan")
    summary.append(["LAPORAN OPERASIONAL LB AUTO"])
    summary["A1"].font = Font(size=18, bold=True, color=RED)
    summary.append(["Dibuat", data["generated_at"]])
    for label, key in (("Total unit", "total_units"), ("Unit terjual", "sold_units"), ("Omzet", "revenue"), ("Total HPP", "hpp"), ("Profit", "profit")):
        summary.append([label, data["summary"][key]])
    for cell in (summary["B5"], summary["B6"], summary["B7"]):
        cell.number_format = '"Rp" #,##0'
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 26

    units_sheet = _sheet(workbook, "Unit", ["ID", "Status", "Merek", "Model", "Tahun", "Warna", "Plat", "Transmisi", "Odometer", "VIN", "Nomor mesin", "Sumber", "Seller", "WhatsApp seller", "Lokasi", "Harga penawaran", "Target harga", "Harga beli", "Biaya repair", "HPP", "Harga jual", "Profit", "Dibuat", "Diperbarui"])
    workflow = _sheet(workbook, "Alur & Approval", ["Unit", "Initial QC", "Catatan Initial QC", "Legal precheck", "Catatan legal", "Checker", "Inspeksi dikirim", "Keputusan HOD", "Voucher", "Pembayaran", "Serah-terima repair", "Repair QC", "Document QC", "Booking", "Finance approved", "Deal", "Delivery dijadwalkan", "Delivery selesai"])
    inspections = _sheet(workbook, "Inspeksi", ["Unit", "Skor body", "Tabrak besar", "Banjir", "Mesin", "Oli", "Kaki-kaki", "Pajak", "Catatan", "Dikirim"])
    repairs = _sheet(workbook, "Pekerjaan Repair", ["Unit", "Kategori", "Panel", "Progres", "Estimasi", "Realisasi", "Vendor", "Stage", "Target", "Catatan"])
    documents = _sheet(workbook, "Dokumen", ["Unit", "STNK", "Jatuh tempo pajak", "Jatuh tempo plat", "BPKB", "Nomor BPKB", "Faktur", "Kwitansi", "KTP pemilik", "Siap jual", "QC pada"])
    sales = _sheet(workbook, "Penjualan", ["Unit", "Customer", "WhatsApp", "NIK", "Alamat", "Skema", "Leasing", "Tenor", "DP", "Harga final", "Status", "Booking", "Finance reference", "Finance approved", "Deal", "Jadwal delivery", "Delivery selesai", "Catatan"])
    photos = _sheet(workbook, "Tautan Foto", ["Unit", "Kategori", "Label", "Link foto"])
    audit_sheet = _sheet(workbook, "Audit Timeline", ["Unit", "Waktu", "Pengguna", "Tindakan", "Entitas", "Detail"])

    for unit in data["units"]:
        units_sheet.append([unit.get(key, "") for key in ("id", "status", "brand", "model", "year", "color", "plate", "transmission", "km", "vin", "engine_number", "source", "seller", "seller_phone", "location", "offer_price", "target_price", "buy_price", "repair_cost")] + [unit["hpp"], unit.get("sell_price", 0), unit["profit"], unit.get("created_at", ""), unit.get("updated_at", "")])
        initial, legal, inspection = unit.get("initial_qc") or {}, unit.get("legal_precheck") or {}, unit.get("inspection") or {}
        payment, repair, docs = unit.get("payment") or {}, unit.get("repair") or {}, unit.get("documents") or {}
        sale = unit.get("sale") or {}
        workflow.append([unit["id"], initial.get("reviewed_at", ""), initial.get("notes", ""), legal.get("approved_at", ""), legal.get("notes", ""), unit.get("assigned_checker_id", ""), inspection.get("submitted_at", ""), unit.get("rejection_reason") or ("Approved" if unit.get("buy_price") else ""), payment.get("voucher_number", ""), payment.get("paid_at", ""), repair.get("handover_at", ""), repair.get("qc_at", ""), docs.get("qc_at", ""), sale.get("booked_at", ""), sale.get("finance_approved_at", ""), sale.get("deal_at", ""), sale.get("delivery_scheduled_at", ""), sale.get("delivery_completed_at", "")])
        if inspection:
            inspections.append([unit["id"], inspection.get("body_score"), "Ya" if inspection.get("major_accident") else "Tidak", "Ya" if inspection.get("flood") else "Tidak", inspection.get("engine_condition"), inspection.get("oil_condition"), inspection.get("suspension_condition"), inspection.get("tax_status"), inspection.get("notes"), inspection.get("submitted_at")])
        items = _list(repair.get("work_items")) or [{}]
        for item in items:
            repairs.append([unit["id"], item.get("category", ""), item.get("panel", ""), item.get("progress", repair.get("progress", "")), item.get("estimated_cost", repair.get("estimated_cost", 0)), item.get("actual_cost", repair.get("actual_cost", 0)), repair.get("vendor", ""), repair.get("stage", ""), repair.get("target_date", ""), repair.get("notes", "")])
        documents.append([unit["id"], docs.get("stnk_status", ""), docs.get("tax_due", ""), docs.get("plate_due", ""), docs.get("bpkb_status", ""), docs.get("bpkb_number", ""), docs.get("invoice_status", ""), "Ada" if docs.get("receipt_available") else "Tidak", "Ada" if docs.get("owner_id_copy") else "Tidak", "Ya" if docs.get("ready_for_sale") else "Tidak", docs.get("qc_at", "")])
        if sale:
            sales.append([unit["id"], sale.get("buyer_name"), sale.get("buyer_phone"), sale.get("buyer_nik"), sale.get("buyer_address"), sale.get("payment_scheme"), sale.get("leasing_vendor"), sale.get("tenor_months"), sale.get("down_payment"), sale.get("final_price"), sale.get("status"), sale.get("booked_at"), sale.get("finance_reference"), sale.get("finance_approved_at"), sale.get("deal_at"), sale.get("delivery_scheduled_at"), sale.get("delivery_completed_at"), sale.get("notes")])
        for category, label, link, _ in unit_photos(unit, base_url):
            photos.append([unit["id"], category, label, link])
            cell = photos.cell(photos.max_row, 4)
            cell.hyperlink = link
            cell.style = "Hyperlink"
        for log in unit.get("audit", []):
            audit_sheet.append([unit["id"], log.get("created_at"), log.get("user_name") or "System", log.get("action"), log.get("entity_type"), json.dumps(log.get("payload") or {}, ensure_ascii=False)])

    _finish_sheet(units_sheet, ("P", "Q", "R", "S", "T", "U", "V"))
    _finish_sheet(workflow)
    _finish_sheet(inspections)
    _finish_sheet(repairs, ("E", "F"))
    _finish_sheet(documents)
    _finish_sheet(sales, ("I", "J"))
    _finish_sheet(photos)
    _finish_sheet(audit_sheet)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _font(size, bold=False):
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


class PdfPages:
    width, height, margin = 1240, 1754, 80

    def __init__(self):
        self.pages = []
        self.page = None
        self.draw = None
        self.y = 0
        self.new_page()

    def new_page(self):
        self.page = Image.new("RGB", (self.width, self.height), "white")
        self.draw = ImageDraw.Draw(self.page)
        self.pages.append(self.page)
        self.draw.rectangle((0, 0, self.width, 18), fill="#EC1D23")
        self.draw.text((self.margin, 40), "LB AUTO  |  LAPORAN OPERASIONAL", font=_font(18, True), fill="#EC1D23")
        self.y = 92

    def ensure(self, height):
        if self.y + height > self.height - 80:
            self.new_page()

    def title(self, text, size=34):
        self.ensure(size + 38)
        self.draw.text((self.margin, self.y), str(text), font=_font(size, True), fill="#17201E")
        self.y += size + 25

    def section(self, text):
        self.ensure(56)
        self.draw.rounded_rectangle((self.margin, self.y, self.width - self.margin, self.y + 42), 8, fill="#F4F5F2")
        self.draw.text((self.margin + 16, self.y + 9), str(text), font=_font(18, True), fill="#EC1D23")
        self.y += 57

    def line(self, label, value):
        value = "-" if value in (None, "") else str(value)
        wrapped = textwrap.wrap(value, width=76) or ["-"]
        line_height = 25 * len(wrapped) + 10
        self.ensure(line_height)
        self.draw.text((self.margin, self.y), str(label), font=_font(15, True), fill="#6F7976")
        for index, part in enumerate(wrapped):
            self.draw.text((390, self.y + index * 25), part, font=_font(15), fill="#17201E")
        self.y += line_height

    def photo(self, label, path, base_dir):
        self.ensure(410)
        x1, x2 = self.margin, self.width - self.margin
        self.draw.rounded_rectangle((x1, self.y, x2, self.y + 360), 12, outline="#E2E4DF", width=2)
        local = None
        raw = str(path or "")
        if raw.startswith("/") and not raw.startswith("//"):
            candidate = (base_dir / raw.lstrip("/")).resolve()
            if base_dir.resolve() in candidate.parents and candidate.is_file():
                local = candidate
        if local:
            try:
                with Image.open(local) as source:
                    image = ImageOps.contain(source.convert("RGB"), (x2 - x1 - 20, 300), Image.Resampling.LANCZOS)
                self.page.paste(image, (x1 + (x2 - x1 - image.width) // 2, self.y + 10))
            except OSError:
                local = None
        if not local:
            self.draw.text((x1 + 24, self.y + 135), "Foto eksternal / tidak tersedia pada server", font=_font(16), fill="#6F7976")
        self.draw.text((x1 + 18, self.y + 326), str(label), font=_font(16, True), fill="#17201E")
        self.y += 380


def _money(value):
    return f"Rp {int(value or 0):,}".replace(",", ".")


def _build_pdf_legacy(data, base_dir, base_url):
    pdf = PdfPages()
    pdf.title("Laporan Lengkap LB AUTO", 42)
    pdf.line("Dibuat", data["generated_at"])
    pdf.section("Ringkasan")
    pdf.line("Total unit", data["summary"]["total_units"])
    pdf.line("Unit terjual", data["summary"]["sold_units"])
    pdf.line("Omzet", _money(data["summary"]["revenue"]))
    pdf.line("Total HPP", _money(data["summary"]["hpp"]))
    pdf.line("Profit", _money(data["summary"]["profit"]))

    for unit in data["units"]:
        pdf.new_page()
        pdf.title(f"{unit['brand']} {unit['model']}")
        pdf.line("ID unit", unit["id"])
        pdf.line("Status akhir", unit["status"])
        pdf.section("Sourcing & Identitas")
        for label, value in (("Sumber", unit.get("source")), ("Seller", unit.get("seller")), ("WhatsApp seller", unit.get("seller_phone")), ("Lokasi", unit.get("location")), ("Tahun", unit.get("year")), ("Warna", unit.get("color")), ("Nomor polisi", unit.get("plate")), ("Transmisi", unit.get("transmission")), ("Odometer", f"{unit.get('km', 0):,} km"), ("VIN", unit.get("vin")), ("Nomor mesin", unit.get("engine_number"))):
            pdf.line(label, value)
        pdf.section("Harga & Profitabilitas")
        for label, value in (("Harga penawaran", _money(unit.get("offer_price"))), ("Target harga", _money(unit.get("target_price"))), ("Harga beli", _money(unit.get("buy_price"))), ("Biaya repair", _money(unit.get("repair_cost"))), ("HPP", _money(unit["hpp"])), ("Harga jual", _money(unit.get("sell_price"))), ("Profit", _money(unit["profit"]))):
            pdf.line(label, value)

        initial, legal, inspection = unit.get("initial_qc") or {}, unit.get("legal_precheck") or {}, unit.get("inspection") or {}
        pdf.section("QC, Legal & Inspeksi")
        for label, value in (("Initial QC", initial.get("reviewed_at")), ("Catatan Initial QC", initial.get("notes")), ("Legal precheck", legal.get("approved_at")), ("Catatan legal", legal.get("notes")), ("Skor body", inspection.get("body_score")), ("Tabrak besar", "Ya" if inspection.get("major_accident") else "Tidak"), ("Banjir", "Ya" if inspection.get("flood") else "Tidak"), ("Mesin", inspection.get("engine_condition")), ("Oli", inspection.get("oil_condition")), ("Kaki-kaki", inspection.get("suspension_condition")), ("Pajak", inspection.get("tax_status")), ("Catatan checker", inspection.get("notes")), ("Laporan dikirim", inspection.get("submitted_at"))):
            pdf.line(label, value)

        payment, repair, docs = unit.get("payment") or {}, unit.get("repair") or {}, unit.get("documents") or {}
        pdf.section("Pembayaran Pembelian")
        for label, value in (("Voucher", payment.get("voucher_number")), ("Nominal", _money(payment.get("amount"))), ("Metode", payment.get("method")), ("Diajukan", payment.get("requested_at")), ("Dibayar", payment.get("paid_at"))):
            pdf.line(label, value)
        pdf.section("Repair")
        for label, value in (("Serah-terima", repair.get("handover_at")), ("Odometer diterima", repair.get("handover_odometer")), ("Vendor", repair.get("vendor")), ("Stage", repair.get("stage")), ("Progres", f"{repair.get('progress', 0)}%"), ("Estimasi", _money(repair.get("estimated_cost"))), ("Realisasi", _money(repair.get("actual_cost"))), ("Target", repair.get("target_date")), ("Repair QC", repair.get("qc_status")), ("Catatan QC", repair.get("qc_notes"))):
            pdf.line(label, value)
        for index, item in enumerate(_list(repair.get("work_items")), 1):
            pdf.line(f"Pekerjaan {index}", f"{item.get('category', '')} — {item.get('panel', '')} — {item.get('progress', 0)}% — estimasi {_money(item.get('estimated_cost'))} — realisasi {_money(item.get('actual_cost'))}")

        pdf.section("Document QC")
        for label, value in (("STNK", docs.get("stnk_status")), ("Pajak", docs.get("tax_due")), ("BPKB", docs.get("bpkb_status")), ("Nomor BPKB", docs.get("bpkb_number")), ("Faktur", docs.get("invoice_status")), ("Kwitansi", "Ada" if docs.get("receipt_available") else "Tidak"), ("KTP pemilik", "Ada" if docs.get("owner_id_copy") else "Tidak"), ("Siap jual", "Ya" if docs.get("ready_for_sale") else "Tidak"), ("QC pada", docs.get("qc_at"))):
            pdf.line(label, value)

        listing, sale, greeting = unit.get("listing") or {}, unit.get("sale") or {}, unit.get("greeting") or {}
        pdf.section("Listing & Penjualan")
        for label, value in (("Harga cash listing", _money(listing.get("cash_price"))), ("Harga kredit listing", _money(listing.get("credit_price"))), ("Channel", ", ".join(_list(listing.get("channels")))), ("Published", listing.get("published_at")), ("Customer", sale.get("buyer_name")), ("WhatsApp", sale.get("buyer_phone")), ("NIK", sale.get("buyer_nik")), ("Alamat", sale.get("buyer_address")), ("Skema", sale.get("payment_scheme")), ("Leasing", sale.get("leasing_vendor")), ("Tenor", sale.get("tenor_months")), ("DP", _money(sale.get("down_payment"))), ("Harga final", _money(sale.get("final_price"))), ("Booking", sale.get("booked_at")), ("Finance reference", sale.get("finance_reference")), ("Finance approved", sale.get("finance_approved_at")), ("Deal", sale.get("deal_at")), ("Jadwal delivery", sale.get("delivery_scheduled_at")), ("Delivery selesai", sale.get("delivery_completed_at")), ("Rating greeting", greeting.get("rating")), ("Catatan greeting", greeting.get("notes"))):
            pdf.line(label, value)

        if unit.get("audit"):
            pdf.section("Audit Timeline")
            for log in unit["audit"]:
                pdf.line(log.get("created_at", ""), f"{log.get('user_name') or 'System'} — {log.get('action')} — {json.dumps(log.get('payload') or {}, ensure_ascii=False)}")

        photo_records = unit_photos(unit, base_url)
        if photo_records:
            pdf.new_page()
            pdf.title(f"Galeri Foto — {unit['id']}", 30)
            for category, label, _, raw_path in photo_records:
                pdf.photo(f"{category} — {label}", raw_path, base_dir)

    output = BytesIO()
    pdf.pages[0].save(output, format="PDF", save_all=True, append_images=pdf.pages[1:], resolution=150.0)
    output.seek(0)
    return output


LANDSCAPE_WIDTH = 1754
LANDSCAPE_HEIGHT = 1240


def _short(draw, value, font, width):
    text = "-" if value in (None, "") else str(value)
    if draw.textlength(text, font=font) <= width:
        return text
    while len(text) > 2 and draw.textlength(text + "…", font=font) > width:
        text = text[:-1]
    return text.rstrip() + "…"


def _local_photo(raw_path, base_dir):
    raw = str(raw_path or "")
    if not raw.startswith("/") or raw.startswith("//"):
        return None
    candidate = (base_dir / raw.lstrip("/")).resolve()
    if base_dir.resolve() not in candidate.parents or not candidate.is_file():
        return None
    try:
        with Image.open(candidate) as source:
            return source.convert("RGB")
    except OSError:
        return None


def _photo_card(page, draw, box, label, raw_path, base_dir, crop=True):
    x, y, width, height = (int(value) for value in box)
    draw.rounded_rectangle((x, y, x + width, y + height), 10, fill="#F4F5F2", outline="#E2E4DF", width=2)
    label_height = 28
    image = _local_photo(raw_path, base_dir)
    if image:
        area = (width - 4, height - label_height - 4)
        rendered = ImageOps.fit(image, area, Image.Resampling.LANCZOS) if crop else ImageOps.contain(image, area, Image.Resampling.LANCZOS)
        paste_x = x + 2 + (area[0] - rendered.width) // 2
        paste_y = y + 2 + (area[1] - rendered.height) // 2
        page.paste(rendered, (paste_x, paste_y))
    else:
        draw.text((x + 14, y + (height - label_height) // 2 - 8), "Foto belum tersedia", font=_font(12), fill="#8B9491")
    draw.rectangle((x + 1, y + height - label_height, x + width - 1, y + height - 1), fill="#FFFFFF")
    draw.text((x + 10, y + height - 22), _short(draw, label, _font(11, True), width - 20), font=_font(11, True), fill="#17201E")


def _section_box(draw, box, title, fields, columns=2):
    x, y, width, height = box
    draw.rounded_rectangle((x, y, x + width, y + height), 12, fill="#FFFFFF", outline="#DFE2DD", width=2)
    draw.rounded_rectangle((x + 1, y + 1, x + width - 1, y + 38), 10, fill="#F4F5F2")
    draw.rectangle((x + 1, y + 27, x + width - 1, y + 39), fill="#F4F5F2")
    draw.text((x + 15, y + 10), title.upper(), font=_font(14, True), fill="#EC1D23")
    rows = max(1, (len(fields) + columns - 1) // columns)
    cell_width = (width - 30) / columns
    row_height = max(30, (height - 50) / rows)
    label_font, value_font = _font(10, True), _font(13)
    for index, (label, value) in enumerate(fields):
        column, row = index % columns, index // columns
        cell_x = x + 15 + column * cell_width
        cell_y = y + 48 + row * row_height
        draw.text((cell_x, cell_y), str(label).upper(), font=label_font, fill="#7A8581")
        draw.text((cell_x, cell_y + 15), _short(draw, value, value_font, cell_width - 18), font=value_font, fill="#17201E")


def _metric_box(draw, box, label, value, primary=False):
    x, y, width, height = box
    fill = "#EC1D23" if primary else "#FFFFFF"
    border = "#EC1D23" if primary else "#DFE2DD"
    draw.rounded_rectangle((x, y, x + width, y + height), 11, fill=fill, outline=border, width=2)
    draw.text((x + 13, y + 12), label.upper(), font=_font(10, True), fill="#FFD9DA" if primary else "#7A8581")
    value_font = _font(13 if len(str(value)) > 14 else 17, True)
    draw.text((x + 13, y + 35), _short(draw, value, value_font, width - 26), font=value_font, fill="#FFFFFF" if primary else "#17201E")


def _page_base(title, subtitle=""):
    page = Image.new("RGB", (LANDSCAPE_WIDTH, LANDSCAPE_HEIGHT), "white")
    draw = ImageDraw.Draw(page)
    draw.rectangle((0, 0, LANDSCAPE_WIDTH, 16), fill="#EC1D23")
    draw.text((55, 34), "LB AUTO", font=_font(18, True), fill="#EC1D23")
    draw.text((55, 61), title, font=_font(32, True), fill="#17201E")
    if subtitle:
        draw.text((55, 101), subtitle, font=_font(13), fill="#6F7976")
    draw.text((LANDSCAPE_WIDTH - 250, 46), "LAPORAN OPERASIONAL", font=_font(13, True), fill="#6F7976")
    return page, draw


def _summary_page(data):
    page, draw = _page_base("Laporan Lengkap LB AUTO", f"Dibuat {data['generated_at']} · Seluruh unit dan perjalanan operasional")
    metrics = [
        ("Total unit", str(data["summary"]["total_units"]), True),
        ("Unit terjual", str(data["summary"]["sold_units"]), False),
        ("Omzet", _money(data["summary"]["revenue"]), False),
        ("Total HPP", _money(data["summary"]["hpp"]), False),
        ("Profit", _money(data["summary"]["profit"]), False),
    ]
    gap, x, y = 14, 55, 145
    width = (LANDSCAPE_WIDTH - 110 - gap * 4) / 5
    for index, (label, value, primary) in enumerate(metrics):
        _metric_box(draw, (x + index * (width + gap), y, width, 86), label, value, primary)

    counts = {}
    for unit in data["units"]:
        counts[unit["status"]] = counts.get(unit["status"], 0) + 1
    status_fields = [(status.replace("_", " "), value) for status, value in sorted(counts.items())]
    _section_box(draw, (55, 255, 520, 420), "Komposisi status", status_fields, 2)

    x1, y1, width1, height1 = 595, 255, 1104, 880
    draw.rounded_rectangle((x1, y1, x1 + width1, y1 + height1), 12, fill="#FFFFFF", outline="#DFE2DD", width=2)
    draw.rounded_rectangle((x1 + 1, y1 + 1, x1 + width1 - 1, y1 + 44), 10, fill="#F4F5F2")
    draw.text((x1 + 16, y1 + 12), "DAFTAR UNIT", font=_font(14, True), fill="#EC1D23")
    headers = [("ID UNIT", 0), ("KENDARAAN", 205), ("STATUS", 525), ("HPP", 745), ("HARGA JUAL", 900)]
    for label, offset in headers:
        draw.text((x1 + 16 + offset, y1 + 61), label, font=_font(10, True), fill="#7A8581")
    row_y = y1 + 88
    for index, unit in enumerate(data["units"][:18]):
        if index % 2:
            draw.rectangle((x1 + 8, row_y - 6, x1 + width1 - 8, row_y + 31), fill="#FAFAF8")
        values = [unit["id"], f"{unit['brand']} {unit['model']} ({unit['year']})", unit["status"].replace("_", " "), _money(unit["hpp"]), _money(unit.get("sell_price"))]
        widths = [180, 300, 200, 140, 155]
        for (value, (_, offset), max_width) in zip(values, headers, widths):
            draw.text((x1 + 16 + offset, row_y), _short(draw, value, _font(12), max_width), font=_font(12), fill="#17201E")
        row_y += 43
    if len(data["units"]) > 18:
        draw.text((x1 + 16, y1 + height1 - 34), f"+ {len(data['units']) - 18} unit lainnya tersedia pada halaman berikutnya", font=_font(11, True), fill="#EC1D23")
    draw.text((55, LANDSCAPE_HEIGHT - 52), "Satu halaman berikutnya mewakili satu unit.", font=_font(12), fill="#6F7976")
    return page


def _unit_page(unit, base_dir, base_url):
    page, draw = _page_base(f"{unit['brand']} {unit['model']}", f"{unit['id']} · {unit['plate']} · {unit['status'].replace('_', ' ')}")
    left_x, left_width = 55, 1040
    right_x, right_width = 1120, 579

    metrics = [
        ("Penawaran", _money(unit.get("offer_price")), False),
        ("Harga beli", _money(unit.get("buy_price")), False),
        ("Biaya repair", _money(unit.get("repair_cost")), False),
        ("HPP", _money(unit["hpp"]), False),
        ("Harga jual", _money(unit.get("sell_price")), False),
        ("Profit", _money(unit["profit"]), True),
    ]
    gap = 9
    metric_width = (left_width - gap * 5) / 6
    for index, (label, value, primary) in enumerate(metrics):
        _metric_box(draw, (left_x + index * (metric_width + gap), 133, metric_width, 72), label, value, primary)

    identity = [
        ("Sumber", unit.get("source")), ("Seller", unit.get("seller")),
        ("WhatsApp seller", unit.get("seller_phone")), ("Lokasi", unit.get("location")),
        ("Tahun / Warna", f"{unit.get('year', '-')} / {unit.get('color', '-') }"), ("Transmisi", unit.get("transmission")),
        ("Odometer", f"{unit.get('km', 0):,} km"), ("VIN", unit.get("vin")),
        ("Nomor mesin", unit.get("engine_number")), ("Dibuat", unit.get("created_at")),
    ]
    _section_box(draw, (left_x, 220, left_width, 178), "Sourcing & identitas", identity, 2)

    initial, legal, inspection = unit.get("initial_qc") or {}, unit.get("legal_precheck") or {}, unit.get("inspection") or {}
    qc_fields = [
        ("Initial QC", initial.get("reviewed_at")), ("Catatan Initial QC", initial.get("notes")),
        ("Legal precheck", legal.get("approved_at")), ("Catatan legal", legal.get("notes")),
        ("Inspeksi dikirim", inspection.get("submitted_at")), ("Skor body", inspection.get("body_score")),
        ("Tabrak besar / Banjir", f"{'Ya' if inspection.get('major_accident') else 'Tidak'} / {'Ya' if inspection.get('flood') else 'Tidak'}"),
        ("Mesin / Oli", f"{inspection.get('engine_condition') or '-'} / {inspection.get('oil_condition') or '-'}"),
        ("Kaki-kaki / Pajak", f"{inspection.get('suspension_condition') or '-'} / {inspection.get('tax_status') or '-'}"),
        ("Catatan checker", inspection.get("notes")),
    ]
    _section_box(draw, (left_x, 414, left_width, 200), "QC, legal & inspeksi", qc_fields, 2)

    payment, repair, docs = unit.get("payment") or {}, unit.get("repair") or {}, unit.get("documents") or {}
    repair_fields = [
        ("Voucher", payment.get("voucher_number")), ("Pembayaran pembelian", payment.get("paid_at")),
        ("Serah-terima repair", repair.get("handover_at")), ("Vendor / Stage", f"{repair.get('vendor') or '-'} / {repair.get('stage') or '-'}"),
        ("Progres repair", f"{repair.get('progress', 0)}%"), ("Realisasi repair", _money(repair.get("actual_cost"))),
        ("Repair QC", f"{repair.get('qc_status') or '-'} · {repair.get('qc_at') or '-'}"), ("Catatan Repair QC", repair.get("qc_notes")),
        ("STNK / BPKB", f"{docs.get('stnk_status') or '-'} / {docs.get('bpkb_status') or '-'}"), ("Faktur", docs.get("invoice_status")),
        ("Document QC", docs.get("qc_at")), ("Siap jual", "Ya" if docs.get("ready_for_sale") else "Tidak"),
    ]
    _section_box(draw, (left_x, 630, left_width, 218), "Pembayaran, repair & dokumen", repair_fields, 2)

    listing, sale, greeting = unit.get("listing") or {}, unit.get("sale") or {}, unit.get("greeting") or {}
    sale_fields = [
        ("Published", listing.get("published_at")), ("Channel", ", ".join(_list(listing.get("channels")))),
        ("Customer", sale.get("buyer_name")), ("WhatsApp customer", sale.get("buyer_phone")),
        ("Skema / Leasing", f"{sale.get('payment_scheme') or '-'} / {sale.get('leasing_vendor') or '-'}"), ("DP / Harga final", f"{_money(sale.get('down_payment'))} / {_money(sale.get('final_price'))}"),
        ("Booking", sale.get("booked_at")), ("Finance approved", sale.get("finance_approved_at")),
        ("Deal", sale.get("deal_at")), ("Jadwal delivery", sale.get("delivery_scheduled_at")),
        ("Delivery selesai", sale.get("delivery_completed_at")), ("Greeting", f"Rating {greeting.get('rating') or '-'} · {greeting.get('notes') or '-'}"),
    ]
    _section_box(draw, (left_x, 864, left_width, 218), "Listing, penjualan & delivery", sale_fields, 2)

    audit_items = unit.get("audit") or []
    audit_text = " · ".join(f"{str(item.get('created_at', ''))[:10]} {item.get('action', '')}" for item in audit_items[-6:]) or "Belum ada aktivitas audit"
    _section_box(draw, (left_x, 1098, left_width, 92), "Audit timeline terakhir", [("Aktivitas", audit_text)], 1)

    draw.text((right_x, 133), "DOKUMENTASI FOTO", font=_font(14, True), fill="#EC1D23")
    source_photos = _list(unit.get("source_photos"))
    _photo_card(page, draw, (right_x, 158, right_width, 215), "Foto cover unit", source_photos[0] if source_photos else None, base_dir)
    draw.text((right_x, 390), "INSPEKSI 8 SISI", font=_font(12, True), fill="#6F7976")
    inspection_photos = _list(inspection.get("photos"))
    photo_gap = 8
    card_width = (right_width - photo_gap * 2) / 3
    card_height = 137
    for index, label in enumerate(PHOTO_LABELS):
        column, row = index % 3, index // 3
        raw = inspection_photos[index] if index < len(inspection_photos) else None
        _photo_card(page, draw, (right_x + column * (card_width + photo_gap), 416 + row * (card_height + photo_gap), card_width, card_height), f"{index + 1:02d} {label}", raw, base_dir)

    all_photos = unit_photos(unit, base_url)
    extras = [item for item in all_photos if item[0] not in ("Cover", "Inspeksi")]
    extras_y = 416 + 3 * (card_height + photo_gap) + 12
    draw.text((right_x, extras_y), "DOKUMENTASI TAMBAHAN", font=_font(12, True), fill="#6F7976")
    extra_height = 116
    for index, (category, label, _, raw) in enumerate(extras[:6]):
        column, row = index % 3, index // 3
        _photo_card(page, draw, (right_x + column * (card_width + photo_gap), extras_y + 25 + row * (extra_height + photo_gap), card_width, extra_height), f"{category} · {label}", raw, base_dir)
    if not extras:
        draw.text((right_x, extras_y + 29), "Belum ada dokumentasi tambahan", font=_font(12), fill="#8B9491")
    elif len(extras) > 6:
        draw.text((right_x, LANDSCAPE_HEIGHT - 35), f"+ {len(extras) - 6} foto lain tersedia melalui tautan pada export Excel", font=_font(10), fill="#6F7976")

    draw.text((55, LANDSCAPE_HEIGHT - 28), "LB AUTO · Laporan perjalanan unit dari sourcing sampai delivery", font=_font(10), fill="#8B9491")
    return page


def build_pdf(data, base_dir, base_url):
    pages = [_summary_page(data)]
    pages.extend(_unit_page(unit, base_dir, base_url) for unit in data["units"])
    output = BytesIO()
    pages[0].save(output, format="PDF", save_all=True, append_images=pages[1:], resolution=150.0, quality=88)
    output.seek(0)
    return output
